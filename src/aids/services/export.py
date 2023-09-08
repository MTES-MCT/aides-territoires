from io import BytesIO
import logging
import os
from openpyxl import Workbook
from openpyxl.utils.escape import escape
from xhtml2pdf import pisa
from datetime import date, datetime, timedelta

from django.http import HttpResponse
from django.conf import settings
from django.contrib.sites.models import Site
from django.template.loader import get_template
from django.utils.text import get_valid_filename
from django.utils import timezone
from django.db.models import Count
from django.core import files
from django.core.files.base import ContentFile
from aids.constants import (
    COLLECTIVITIES_AUDIENCES,
    FINANCIAL_AIDS_LIST,
    TECHNICAL_AIDS_LIST,
)

from aids.resources import AidResourcePublic
from aids.models import Aid, AidProject
from accounts.models import User
from core.utils import get_base_url
from organizations.models import Organization
from stats.models import (
    AidApplicationUrlClickEvent,
    AidOriginUrlClickEvent,
    AidViewEvent,
)
from exporting.models import DataExport

logger = logging.getLogger("console_log")


def fetch_resources(uri: str, rel) -> str:
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access those
    resources
    """
    static_url = settings.STATIC_URL  # Typically /static/
    static_root = settings.STATIC_ROOT  # Typically /some/path/project_static/
    media_url = settings.MEDIA_URL  # Typically /media/
    media_root = settings.MEDIA_ROOT  # Typically /some/path/project_static/media/

    if uri.startswith(media_url):
        path = os.path.join(media_root, uri.replace(media_url, ""))
    elif uri.startswith(static_url):
        path = os.path.join(static_root, uri.replace(static_url, ""))
    else:
        return uri

    # make sure that file exists
    if not os.path.isfile(path):
        raise Exception("media URI must start with %s or %s" % (static_url, media_url))

    return path


def export_aids(organization, file_format: str) -> dict:
    organization = Organization.objects.get(pk=organization)
    users = organization.beneficiaries.values_list("pk", flat=True)
    aids_qs = Aid.objects.live().filter(author__in=users).order_by("pk")

    exported_aids = AidResourcePublic().export(aids_qs)

    today = date.today()
    today_formated = today.strftime("%Y-%m-%d")

    filename = get_valid_filename(
        f"Aides-territoires - {today_formated } - {organization.name}.{file_format}"
    )

    if file_format == "csv":
        response = {
            "content": exported_aids.csv.encode(),
            "content_type": "text/csv",
            "filename": filename,
        }
    elif file_format == "xlsx":
        exported_aids.title = "Aides-territoires"  # Sheet title
        response = {
            "content": exported_aids.xlsx,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": filename,
        }
    elif file_format == "pdf":
        template = get_template("aids/aids_export_pdf.html")

        current_site = Site.objects.get_current()
        html = template.render(
            {
                "today": today,
                "organization": organization,
                "aid_set": aids_qs,
                "hostname": f"https://{current_site.domain}",
            }
        )

        result = BytesIO()
        pdf = pisa.CreatePDF(
            BytesIO(html.encode("utf-8")), dest=result, link_callback=fetch_resources
        )

        if not pdf.err:
            response = {
                "content": result.getvalue(),
                "content_type": "application/pdf",
                "filename": filename,
            }
        else:
            response = {"error": "PDF generation error", "error_detail": pdf.err}

        result.close()
    else:
        response = {"error": "Unknown export format", "error_detail": file_format}

    return response


def export_aid_detail_pdf(aid, user, organization):
    today = date.today()
    today_formated = today.strftime("%Y-%m-%d")

    filename = get_valid_filename(
        f"Aides-territoires - {today_formated } - {aid.name}.pdf"
    )

    template = get_template("aids/aid_export_pdf.html")

    current_site = Site.objects.get_current()
    html = template.render(
        {
            "today": today,
            "aid": aid,
            "organization": organization,
            "hostname": f"https://{current_site.domain}",
        }
    )

    result = BytesIO()
    pdf = pisa.CreatePDF(
        BytesIO(html.encode("utf-8")), dest=result, link_callback=fetch_resources
    )

    if not pdf.err:
        response = {
            "content": result.getvalue(),
            "content_type": "application/pdf",
            "filename": filename,
        }
    else:
        response = {"error": "PDF generation error", "error_detail": pdf.err}

    result.close()

    return response


def date_range_list(start_date, end_date):
    # Return list of datetime.date objects (inclusive) between start_date and end_date (inclusive).
    date_list = []
    curr_date = start_date
    while curr_date <= end_date:
        date_list.append(curr_date)
        curr_date += timedelta(days=1)
    return date_list


def export_aid_stats(
    aid,
    start_date,
    end_date,
) -> dict:
    if end_date is None:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()

    date_list = date_range_list(start_date, end_date)

    aid_name = aid.name

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=Aides-territoires-statistiques.xlsx"

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = f"Aides-territoires-statistiques-{aid_name}"

    # Define the titles for columns
    columns = [
        "Date",
        "Nombre de vues",
        "Nombre de clics sur Candidater",
        "Nombre de clics sur Plus d’informations",
        "Nombre de projets privés liés",
        "Nombre de projets publics liés",
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Define the data for each cell in the row
    aidviewevent_qs = (
        AidViewEvent.objects.filter(aid=aid.id)
        .extra({"date_created": "date(date_created)"})
        .values("date_created")
        .annotate(view_count=Count("date_created", distinct=True))
    )

    aidapplicationurlclickevent_qs = (
        AidApplicationUrlClickEvent.objects.filter(aid=aid.id)
        .extra({"date_created": "date(date_created)"})
        .values("date_created")
        .annotate(click_count=Count("date_created", distinct=True))
    )

    aidoriginurlclickevent_qs = (
        AidOriginUrlClickEvent.objects.filter(aid=aid.id)
        .extra({"date_created": "date(date_created)"})
        .values("date_created")
        .annotate(click_count=Count("date_created", distinct=True))
    )

    private_projects_linked_qs = (
        AidProject.objects.filter(aid=aid.id, project__is_public=False)
        .extra({"date_created": "date(aids_aidproject.date_created)"})
        .values("date_created")
        .annotate(linked_count=Count("date_created", distinct=True))
    )

    public_projects_linked_qs = (
        AidProject.objects.filter(aid=aid.id, project__is_public=True)
        .extra({"date_created": "date(aids_aidproject.date_created)"})
        .values("date_created")
        .annotate(linked_count=Count("date_created", distinct=True))
    )

    for day in date_list:
        row_num += 1
        row = [
            day.strftime("%d-%m-%Y"),
        ]
        day = datetime.strftime(day, "%Y-%m-%d")

        viewevent = 0
        for x in aidviewevent_qs:
            if datetime.strftime(x["date_created"], "%Y-%m-%d") == day:
                viewevent = x["view_count"]
        row.append(viewevent)

        aidapplicationurlclickevent = 0
        for x in aidapplicationurlclickevent_qs:
            if datetime.strftime(x["date_created"], "%Y-%m-%d") == day:
                aidapplicationurlclickevent = x["click_count"]
        row.append(aidapplicationurlclickevent)

        aidoriginurlclickevent = 0
        for x in aidoriginurlclickevent_qs:
            if datetime.strftime(x["date_created"], "%Y-%m-%d") == day:
                aidoriginurlclickevent = x["click_count"]
        row.append(aidoriginurlclickevent)

        private_projects_linked = 0
        for x in private_projects_linked_qs:
            if datetime.strftime(x["date_created"], "%Y-%m-%d") == day:
                private_projects_linked = x["linked_count"]
        row.append(private_projects_linked)

        public_projects_linked = 0
        for x in public_projects_linked_qs:
            if datetime.strftime(x["date_created"], "%Y-%m-%d") == day:
                public_projects_linked = x["linked_count"]
        row.append(public_projects_linked)

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def export_related_projects(aid_id, user_id):
    aid = Aid.objects.get(id=aid_id)
    user = User.objects.get(id=user_id)
    aid_name = aid.name

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = "Projets-ajoutés"

    # Define the titles for columns
    columns = [
        "Nom du projet",
        "Caractère public du projet",
        "Porteur du projet",
        "Périmètre du porteur",
        "Code Insee du périmètre",
        "Type de porteur",
        "Personne ayant ajouté l'aide",
        "Fonction de la personne",
        "email de la personne",
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for project in aid.projects.all():
        row_num += 1
        row = [
            project.name,
        ]

        project_is_public = project.is_public
        if project_is_public is not False:
            project_is_public = "Oui"
        else:
            project_is_public = "Non"
        row.append(project_is_public)

        row.append(project.organization.name)

        if project.organization.perimeter is not None:
            row.append(project.organization.perimeter.name)
            row.append(project.organization.perimeter.insee)
        else:
            row.append("périmètre non communiqué")
            row.append("périmètre non communiqué")

        choices_dict = dict(Organization.ORGANIZATION_TYPE_CHOICES)
        key = project.organization.organization_type[0]
        organization_type_value = choices_dict.get(key, "")
        row.append(organization_type_value)

        aidproject = AidProject.objects.get(aid=aid.id, project=project.id)
        if aidproject.project.author.all() is not None:
            creator = aidproject.project.author.all().first()
            row.append(creator.full_name)

            choices_dict = dict(User.FUNCTION_TYPE)
            key = creator.beneficiary_function
            beneficiary_function_value = choices_dict.get(key, "")
            row.append(beneficiary_function_value)

            row.append(creator.email)
        else:
            row.append("données inconnues")
            row.append("données inconnues")
            row.append("données inconnues")

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # kept the workbook as bytes in an in-memory buffer
    vworkbook = BytesIO()
    workbook.save(vworkbook)

    # takes the value from the Buffer
    content = vworkbook.getvalue()

    content_file = ContentFile(content)
    file_object = files.File(content_file, name=f"Aides-territoires-{aid_name}.xlsx")
    dataexport_object = DataExport.objects.create(
        author_id=user.id,
        exported_file=file_object,
    )
    return dataexport_object


def export_aids_for_collectivities():  # NOSONAR
    user = User.objects.get(email=settings.SERVER_EMAIL)
    coll_audiences = [x for x, y in COLLECTIVITIES_AUDIENCES]
    aids = Aid.objects.filter(
        targeted_audiences__overlap=coll_audiences,
        date_created__year__gte=2019,
        date_created__year__lte=2022,
    )

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = "Aides-collectivites"

    # Define the titles for columns
    columns = [
        "Nom de l’aide",
        "Année de l’aide",
        "Aide périmée",
        "Porteur de l’aide",
        "Catégorie du porteur",
        "Public bénéficiaire",
        "Périmètre de l’aide",
        "Aide en ingénierie",
        "Aide financière",
        "Type de dépenses couvertes",
        "Thématiques de l’aide",
        "Programme de l’aide",
        "Description de l’aide",
        "URL de l’aide",
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for aid in aids:
        row_num += 1
        row = [
            aid.name,
            aid.date_created.year,
        ]

        if aid.has_expired():
            row.append("Oui")
        else:
            row.append("Non")

        financers = aid.financers.all().values_list("name", flat=True)
        if financers:
            row.append(", ".join(financers))
        else:
            row.append("")

        financers_categories = aid.financers.all().values_list(
            "group__subcategory__category__name", flat=True
        )
        if financers_categories and list(financers_categories) != [None]:
            row.append(", ".join(financers_categories))
        else:
            row.append("")

        if aid.targeted_audiences:
            row.append(", ".join(aid.targeted_audiences))
        else:
            row.append("")

        if aid.perimeter:
            row.append(aid.perimeter.name)
        else:
            row.append("")

        if aid.aid_types and set(aid.aid_types).isdisjoint(TECHNICAL_AIDS_LIST):
            row.append("Non")
        else:
            row.append("Oui")

        if aid.aid_types and set(aid.aid_types).isdisjoint(FINANCIAL_AIDS_LIST):
            row.append("Non")
        else:
            row.append("Oui")

        if aid.destinations:
            row.append(", ".join([dict(Aid.DESTINATIONS)[x] for x in aid.destinations]))
        else:
            row.append("")

        row.append(", ".join(aid.categories.values_list("name", flat=True).distinct()))

        row.append(", ".join(aid.programs.values_list("name", flat=True).distinct()))

        row.append(escape(aid.description))

        base_url = get_base_url()
        row.append(f"{base_url}{aid.get_absolute_url()}")

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # kept the workbook as bytes in an in-memory buffer
    vworkbook = BytesIO()
    workbook.save(vworkbook)

    # takes the value from the Buffer
    content = vworkbook.getvalue()

    content_file = ContentFile(content)
    file_object = files.File(
        content_file, name="Aides-territoires-aids-collectivities.xlsx"
    )
    dataexport_object = DataExport.objects.create(
        author_id=user.id,
        exported_file=file_object,
    )

    logger.info(f"{aids.count()} aids exported")
    return dataexport_object
