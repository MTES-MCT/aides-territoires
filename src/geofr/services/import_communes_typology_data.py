"""
This script imports data about if a commune is
rural or urban.

Data source: https://www.insee.fr/fr/statistiques/5039991?sommaire=5040030
"""

from openpyxl import load_workbook
from io import BytesIO
import requests

from geofr.models import Perimeter


def import_communes_rurality_data() -> int:
    base_url = "https://www.insee.fr/fr/statistiques/"
    file_url = "fichier/5039991/FET2021-D4.xlsx"
    response = requests.get(f"{base_url}{file_url}")
    data_file = BytesIO(response.content)

    wb = load_workbook(data_file)
    ws = wb["Figure 5"]

    counter = 0

    for commune_row in ws.iter_rows(values_only=True, min_row=4):
        commune_code = commune_row[0]
        commune_type = commune_row[1]
        if isinstance(commune_code, str) and commune_code.isdigit():
            commune = Perimeter.objects.filter(
                scale=Perimeter.SCALES.commune, is_obsolete=False, code=commune_code
            ).first()

            if commune:
                commune.density_typology = commune_type
                commune.save()
                counter += 1

    return counter
