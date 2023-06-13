import logging
from openpyxl import load_workbook
import requests

from django.db import transaction
from django.utils import timezone
from django.utils.text import BytesIO

from geofr.models import Perimeter
from geofr.utils import attach_perimeters

logger = logging.getLogger("console_log")

"""
Our data source comes from the Ministère de la Cohésion des territoires.
https://www.data.gouv.fr/fr/datasets/schema-de-coherence-territoriale-scot-donnees-sudocuh-dernier-etat-des-lieux-annuel-au-31-decembre-2022/
"""
DATA_URL = "https://www.data.gouv.fr/fr/datasets/r/ffb89961-155d-4134-8682-b10f7aa715c2"
SHEET_NAME = "Communes Scot"


@transaction.atomic
def populate_scots() -> None:
    """
    Import the list of SCoTs.
    """
    start_time = timezone.now()

    scots = {}
    nb_created = 0
    nb_updated = 0
    nb_obsolete = 0

    logger.debug(f"Getting file {DATA_URL} from distant URL")
    response = requests.get(DATA_URL)
    data_file = BytesIO(response.content)

    wb = load_workbook(data_file)
    ws = wb[SHEET_NAME]

    logger.debug("Parsing file…")
    for row in ws.iter_rows(values_only=True, min_row=5):
        scot_id = row[0]
        scot_name = row[1]
        insee_code = row[3]

        # Exclude the credit line at the end ("Réalisation : [...]")
        if isinstance(scot_id, int):
            if scot_id not in scots:
                scots[scot_id] = {"name": scot_name, "communes": []}

            scots[scot_id]["communes"].append(insee_code)

    logger.debug("Importing data for SCoTs…")
    for scot_id in scots.keys():
        # id is just an integer, we use a custom code to make it unique
        scot_code = f"SCOT-{scot_id}"
        scot_name = scots[scot_id]["name"]
        logger.debug(f"Importing data for {scot_code} ({scot_name})")

        # Create or update the SCoT perimeter
        scot, created = Perimeter.objects.update_or_create(
            scale=Perimeter.SCALES.adhoc,
            code=scot_code,
            defaults={"name": scot_name, "is_obsolete": False, "date_obsolete": None},
        )
        if created:
            nb_created += 1
        else:
            nb_updated += 1

        # Link the scot with the related communes
        codes = scots[scot_id]["communes"]
        attach_perimeters(scot, codes)

    # Mark obsolete SCoTs
    nb_obsolete = Perimeter.objects.filter(
        scale=Perimeter.SCALES.adhoc,
        code__startswith="SCOT-",
        date_updated__lt=start_time,
    ).update(is_obsolete=True, date_obsolete=start_time)

    end_time = timezone.now()
    logger.info(f"Import made in {end_time - start_time}.")
    logger.info(f"* {nb_created} entries created")
    logger.info(f"* {nb_updated} entries updated")
    logger.info(f"* {nb_obsolete} entries marked as obsolete")
